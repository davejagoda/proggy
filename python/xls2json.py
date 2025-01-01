#!/usr/bin/env python3

import argparse
import collections
import datetime
import json
import os
import sys

import openpyxl


def process_tab(tab_name, verbose):
    ws = collections.OrderedDict()
    for row in wb[tab_name].rows:
        for cell in row:
            if cell.value is None:
                continue
            if verbose > 0:
                print("{}:{}".format(cell.coordinate, cell.value))
            ws[cell.coordinate] = cell.value
    return ws


parser = argparse.ArgumentParser()
parser.add_argument("xlsfile", help="the source spreadsheet")
parser.add_argument("-o", "--outfile", action="store_true")
parser.add_argument("-v", "--verbose", action="count", default=0)
args = parser.parse_args()

wb = openpyxl.load_workbook(filename=args.xlsfile, read_only=True, data_only=True)

jobj = collections.OrderedDict()
for tab_name in wb.sheetnames:
    if args.verbose > 0:
        print(tab_name)
    jobj[tab_name] = process_tab(tab_name, args.verbose)

output = json.dumps(
    jobj,
    ensure_ascii=False,
    indent=2,
    default=lambda x: (
        x.isoformat()
        if isinstance(x, datetime.datetime)
        else json.JSONEncoder().default(x)
    ),
)

if args.outfile:
    (root, ext) = os.path.splitext(args.xlsfile)
    jsonfile = "{}.json".format(root)
    if os.path.exists(jsonfile):
        print("file {} already exists, refusing to overwrite".format(jsonfile))
        sys.exit(1)
    with open(jsonfile, "w") as f:
        f.write("{}\n".format(output))
else:
    print(output)
